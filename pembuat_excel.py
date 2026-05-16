import pandas as pd
import re
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def buat_laporan_excel():
    nama_file_log = "log_error.txt"
    nama_file_excel = "Analisis_Bedah_Log_Otomatis.xlsx"

    if not os.path.exists(nama_file_log):
        print("File log_error.txt tidak ditemukan")
        return

    with open(nama_file_log, "r") as file:
        baris_log = file.readlines()

    hasil_analisis = []
    nomor = 1

    for baris in baris_log:
        cocok = re.search(r'\[(.+)\] DITOLAK Skor (\d+)% Nominal (.+) Tagihan (.+) Data (\{.+\})', baris)
        if cocok:
            waktu = cocok.group(1)
            skor = int(cocok.group(2))
            nominal = cocok.group(3)
            tagihan = cocok.group(4)
            data_str = cocok.group(5)

            id_trx = "TG-" + waktu.replace("-", "").replace(":", "").replace(" ", "")[-6:]

            if tagihan == "Rp0":
                penyebab = "Variabel tagihan bernilai nol"
                analisis = "Pada transaksi ke-" + str(nomor) + ", aplikasi menolak gambar karena nominal tagihan dari web bernilai nol. Sistem membandingkan angka OCR dengan nol."
                mitigasi = "Membuang data ke folder Review Manual. Programmer harus menambahkan variabel harga pada skrip uji coba masal."
            elif skor == 0:
                penyebab = "Gambar tidak terbaca"
                analisis = "Pada transaksi ke-" + str(nomor) + ", aplikasi melambat karena mesin OCR gagal menemukan teks. Gambar terlalu gelap atau resolusi rusak."
                mitigasi = "Membuang data ke folder Review Manual. Sistem menerapkan penyesuaian kontras gambar otomatis."
            elif "Iklan oleh Google" in data_str or "Detail Transaksi" in data_str or "Lokasi Merchant" in data_str:
                penyebab = "Polusi teks elemen antarmuka"
                analisis = "Pada transaksi ke-" + str(nomor) + ", aplikasi melambat karena mesin OCR mencatat teks tombol layar sebagai nama pengirim dan penerima."
                mitigasi = "Membuang data ke folder Review Manual. Sistem memperbarui aturan kata kunci untuk menghapus frasa aplikasi umum."
            elif skor > 0 and skor < 80:
                penyebab = "Skor gagal mencapai batas minimal"
                analisis = "Pada transaksi ke-" + str(nomor) + ", aplikasi melambat karena skor akurasi hanya mencapai " + str(skor) + " persen akibat rotasi atau teks buram."
                mitigasi = "Membuang data ke folder Review Manual. Penyesuaian batas minimal kelulusan menjadi 60 persen."
            else:
                penyebab = "Gagal validasi otomatis"
                analisis = "Pada transaksi ke-" + str(nomor) + ", aplikasi menolak data akibat informasi yang kurang lengkap."
                mitigasi = "Membuang data ke folder Review Manual."

            hasil_analisis.append({
                "DATA": "DATA " + str(nomor),
                "ID Transaksi": id_trx,
                "Status": "Gagal",
                "Skor Kepercayaan": str(skor) + " persen",
                "Penyebab": penyebab,
                "Analisis": analisis,
                "Aksi Mitigasi": mitigasi
            })
            nomor += 1

    if len(hasil_analisis) == 0:
        print("Tidak ada data log yang sesuai format")
        return

    df = pd.DataFrame(hasil_analisis)
    
    penulis = pd.ExcelWriter(nama_file_excel, engine="openpyxl")
    df.to_excel(penulis, index=False, sheet_name="Tabel Bedah Log")
    
    buku = penulis.book
    lembar = penulis.sheets["Tabel Bedah Log"]
    
    isi_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_header = Font(color="FFFFFF")
    batas = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    
    for kolom in range(1, len(df.columns) + 1):
        sel = lembar.cell(row=1, column=kolom)
        sel.fill = isi_header
        sel.font = font_header
        sel.alignment = Alignment(horizontal="center", vertical="center")
        
    for baris in range(2, len(hasil_analisis) + 2):
        if baris % 2 == 0:
            isi = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        else:
            isi = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
        for kolom in range(1, len(df.columns) + 1):
            sel = lembar.cell(row=baris, column=kolom)
            sel.fill = isi
            sel.border = batas
            sel.alignment = Alignment(wrap_text=True, vertical="top")
            
    lembar.column_dimensions["A"].width = 10
    lembar.column_dimensions["B"].width = 15
    lembar.column_dimensions["C"].width = 10
    lembar.column_dimensions["D"].width = 18
    lembar.column_dimensions["E"].width = 30
    lembar.column_dimensions["F"].width = 50
    lembar.column_dimensions["G"].width = 40
    
    penulis.close()
    print("File Excel berhasil dibuat dengan nama Analisis_Bedah_Log_Otomatis.xlsx")

if __name__ == "__main__":
    buat_laporan_excel()